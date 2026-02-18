#!/usr/bin/env python3
"""
run_claude.py — DSL runner for .prompt files.

Parses and executes .prompt files that batch-process data through Claude.

Usage:
    python3 run_claude.py <file.prompt>
    python3 run_claude.py <file.prompt> --model claude-opus-4-5
    python3 run_claude.py <file.prompt> --dry-run
    python3 run_claude.py <file.prompt> --debug
    python3 run_claude.py <file.prompt> --docx
    python3 run_claude.py <file.prompt> --timeout 120
"""

import argparse
import csv
import json
import os
import re
import subprocess
import sys
from datetime import date
from pathlib import Path


# ── Front-matter parser ───────────────────────────────────────────────────────

def parse_front_matter(text):
    """
    Parse simple YAML-like front matter from the top of a .prompt file.

    Format: key-value lines at the top, terminated by a bare '---' line.
    Supports a single level of nesting for the 'vars:' block.

    Returns (dict, remaining_body_str).
    """
    lines = text.split('\n')
    fm = {}
    vars_dict = {}
    in_vars = False
    i = 0

    while i < len(lines):
        line = lines[i]

        if line.strip() == '---':
            i += 1
            break

        if not line.strip():
            i += 1
            continue

        if line.strip() == 'vars:':
            in_vars = True
            i += 1
            continue

        if in_vars:
            indent_match = re.match(r'^\s{2,}(\w+):\s*(.+)$', line)
            if indent_match:
                key = indent_match.group(1)
                val = indent_match.group(2).strip()
                if val.lower() == 'true':
                    val = True
                elif val.lower() == 'false':
                    val = False
                vars_dict[key] = val
                i += 1
                continue
            else:
                in_vars = False

        kv_match = re.match(r'^(\w+):\s*(.*)$', line)
        if kv_match:
            fm[kv_match.group(1)] = kv_match.group(2).strip()

        i += 1

    if vars_dict:
        fm['vars'] = vars_dict

    body = '\n'.join(lines[i:])
    return fm, body


# ── Prompt file parser ────────────────────────────────────────────────────────

class PromptFile:
    """Parsed representation of a .prompt file."""

    def __init__(self, path):
        self.path = Path(path)
        self.base_dir = self.path.parent
        self.fm = {}
        self.system = None
        self.steps = []           # list of (name, template) for @step blocks
        self.foreach = None       # template for @foreach iteration
        self.output_directive = None  # raw @output...@end block text
        self._parse()

    def _parse(self):
        text = self.path.read_text(encoding='utf-8')

        # Detect and split front matter (present when file begins with "key: value")
        if re.match(r'^\w+:', text) and re.search(r'^---\s*$', text, re.MULTILINE):
            self.fm, body = parse_front_matter(text)
        else:
            body = text

        body = self._resolve_includes(body)
        self._parse_body(body)

    def _resolve_includes(self, text):
        """Replace @include <file> directives with file contents."""
        def replace_include(m):
            include_path = self.base_dir / m.group(1).strip()
            if not include_path.exists():
                print(f"Warning: Include file not found: {include_path}", file=sys.stderr)
                return ''
            return include_path.read_text(encoding='utf-8')

        return re.sub(r'^@include\s+(.+)$', replace_include, text, flags=re.MULTILINE)

    def _parse_body(self, body):
        """Extract directives from body text."""
        # @system block
        sys_match = re.search(r'@system\s*\n(.*?)@end', body, re.DOTALL)
        if sys_match:
            self.system = sys_match.group(1).strip()

        # @output block (keep full text for appending to prompts)
        out_match = re.search(r'@output(?:\s+\w+)?\s*\n.*?@end', body, re.DOTALL)
        if out_match:
            self.output_directive = out_match.group(0)

        # @step blocks
        for m in re.finditer(r'@step\s+(\w+)\s*\n(.*?)@end', body, re.DOTALL):
            self.steps.append((m.group(1), m.group(2).strip()))

        # @foreach block
        foreach_match = re.search(r'@foreach\s*\n(.*?)@end', body, re.DOTALL)
        if foreach_match:
            self.foreach = foreach_match.group(1).strip()

        # Simple prompt: no foreach, no steps — body text (minus special blocks) is the prompt
        if not self.steps and not self.foreach:
            clean = re.sub(r'@system\s*\n.*?@end\n?', '', body, flags=re.DOTALL)
            clean = re.sub(r'@output(?:\s+\w+)?\s*\n.*?@end\n?', '', clean, flags=re.DOTALL)
            clean = clean.strip()
            if clean:
                self.steps = [('main', clean)]


# ── Variable interpolation ────────────────────────────────────────────────────

def interpolate(template, ctx):
    """Replace {{var}} placeholders with values from context dict."""
    def replace(m):
        key = m.group(1).strip()
        return str(ctx[key]) if key in ctx else m.group(0)

    return re.sub(r'\{\{([^}]+)\}\}', replace, template)


def evaluate_condition(cond, ctx):
    """Evaluate simple {{var}} == value or {{var}} != value conditions."""
    m = re.match(r'\{\{(\w+)\}\}\s*(==|!=)\s*(.+)', cond.strip())
    if not m:
        return False
    key, op, expected = m.group(1), m.group(2), m.group(3).strip().strip('"\'')
    actual = str(ctx.get(key, ''))
    return (actual == expected) if op == '==' else (actual != expected)


def process_conditionals(template, ctx):
    """Process @if / @else / @endif blocks."""
    lines = template.split('\n')
    result = []
    i = 0

    while i < len(lines):
        line = lines[i]
        if_match = re.match(r'^@if\s+(.+)$', line)

        if if_match:
            cond_true = evaluate_condition(if_match.group(1), ctx)
            if_lines, else_lines = [], []
            in_else = False
            depth = 1
            i += 1

            while i < len(lines):
                l = lines[i]
                if re.match(r'^@if\s+', l):
                    depth += 1
                    (else_lines if in_else else if_lines).append(l)
                elif re.match(r'^@endif', l):
                    depth -= 1
                    if depth == 0:
                        break
                    (else_lines if in_else else if_lines).append(l)
                elif re.match(r'^@else', l) and depth == 1:
                    in_else = True
                else:
                    (else_lines if in_else else if_lines).append(l)
                i += 1

            result.extend(if_lines if cond_true else else_lines)
        else:
            result.append(line)

        i += 1

    return '\n'.join(result)


def resolve_attachments(template, base_dir):
    """Replace @attach <file> directives with file contents."""
    def replace_attach(m):
        p = base_dir / m.group(1).strip()
        if not p.exists():
            return f"[Attachment not found: {m.group(1)}]"
        return f"\n--- {m.group(1)} ---\n{p.read_text(encoding='utf-8')}\n---\n"

    return re.sub(r'^@attach\s+(.+)$', replace_attach, template, flags=re.MULTILINE)


# ── Data loading ──────────────────────────────────────────────────────────────

def load_data(data_path, sheet=None):
    """
    Load data from JSON, CSV, or XLSX.

    Returns a list of (key, value, row_dict) tuples where:
      key   — identifier for the item (first column or array index)
      value — primary content for {{VALUE}}
      row   — full dict of all fields for interpolation
    """
    p = Path(data_path)
    if not p.exists():
        print(f"[ERROR] Data file not found: {data_path}", file=sys.stderr)
        sys.exit(1)

    suffix = p.suffix.lower()
    if suffix == '.json':
        return _load_json(p)
    elif suffix == '.csv':
        return _load_csv(p)
    elif suffix in ('.xlsx', '.xls'):
        return _load_xlsx(p, sheet)
    else:
        print(f"[ERROR] Unsupported data format: {suffix}", file=sys.stderr)
        sys.exit(1)


def _load_json(path):
    with open(path, encoding='utf-8') as f:
        data = json.load(f)

    items = []
    if isinstance(data, list):
        for i, item in enumerate(data):
            if isinstance(item, dict):
                key = str(item.get(list(item.keys())[0], i))
                value = json.dumps(item)
                row = {str(k): str(v) for k, v in item.items()}
            else:
                key = str(i)
                value = str(item)
                row = {}
            row.update({'KEY': key, 'VALUE': value, 'INDEX': str(i + 1)})
            items.append((key, value, row))
    elif isinstance(data, dict):
        for i, (k, v) in enumerate(data.items()):
            value = v if isinstance(v, str) else json.dumps(v)
            row = {'KEY': k, 'VALUE': value, 'INDEX': str(i + 1)}
            if isinstance(v, dict):
                row.update({str(ck): str(cv) for ck, cv in v.items()})
            items.append((k, value, row))

    return items


def _load_csv(path):
    items = []
    with open(path, encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            row = {k: (v or '') for k, v in row.items()}
            cols = list(row.keys())
            key = row[cols[0]] if cols else str(i)
            # Look for a description/value column, else stringify the whole row
            value = row.get('description', row.get('VALUE', row.get('Description',
                    json.dumps(row))))
            row['KEY'] = key
            row['VALUE'] = value
            row['INDEX'] = str(i + 1)
            items.append((key, value, row))
    return items


def _load_xlsx(path, sheet=None):
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl required. Install: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h is not None else f'col{i}' for i, h in enumerate(rows[0])]
    items = []
    for i, row_vals in enumerate(rows[1:]):
        row = {headers[j]: (str(v) if v is not None else '') for j, v in enumerate(row_vals)}
        key = row.get(headers[0], str(i))
        value = row.get('description', row.get('VALUE', row.get('Description',
                json.dumps(row))))
        row['KEY'] = key
        row['VALUE'] = value
        row['INDEX'] = str(i + 1)
        items.append((key, value, row))

    return items


# ── Claude invocation ─────────────────────────────────────────────────────────

def call_claude(prompt, system=None, model=None, timeout=120, debug=False, dry_run=False):
    """Call the claude CLI and return its text output."""
    if dry_run:
        sep = '=' * 60
        print(f"\n{sep}")
        print("DRY RUN — prompt preview:")
        if system:
            print(f"[SYSTEM] {system[:300]}")
        preview = prompt[:600] + ('...' if len(prompt) > 600 else '')
        print(f"[USER]   {preview}")
        print(f"{sep}\n")
        return "[dry-run output]"

    cmd = ['claude', '-p']
    if system:
        cmd += ['--system', system]
    if model:
        cmd += ['--model', model]

    if debug:
        print(f"[DEBUG] cmd: {' '.join(cmd[:4])}...", file=sys.stderr)
        print(f"[DEBUG] prompt ({len(prompt)} chars): {prompt[:300]}", file=sys.stderr)

    try:
        result = subprocess.run(
            cmd,
            input=prompt,
            capture_output=True,
            text=True,
            timeout=timeout,
            encoding='utf-8',
        )
        if result.returncode != 0 and debug:
            print(f"[WARN] claude exit {result.returncode}: {result.stderr[:200]}", file=sys.stderr)

        output = result.stdout.strip()

        if debug:
            print(f"[DEBUG] response ({len(output)} chars): {output[:300]}", file=sys.stderr)

        return output

    except subprocess.TimeoutExpired:
        print(f"[ERROR] Claude timed out after {timeout}s", file=sys.stderr)
        return ""
    except FileNotFoundError:
        print("[ERROR] 'claude' CLI not found. Install Claude Code: https://claude.ai/download",
              file=sys.stderr)
        sys.exit(1)


# ── Output collection ─────────────────────────────────────────────────────────

def write_output(results, output_path, collect, use_docx):
    """Write collected results to file(s)."""
    if not output_path:
        for _, content in results:
            print(content)
        return

    out = Path(output_path)

    if collect == 'per_item':
        out.mkdir(parents=True, exist_ok=True)
        for key, content in results:
            safe_key = re.sub(r'[^\w\-.]', '_', str(key))[:60]
            item_path = out / f"{safe_key}.txt"
            item_path.write_text(content, encoding='utf-8')
            print(f"  Wrote: {item_path}")
        return

    out.parent.mkdir(parents=True, exist_ok=True)

    if collect == 'append_json':
        parsed = []
        for _, content in results:
            cleaned = re.sub(r'^```(?:json)?\s*\n?', '', content.strip())
            cleaned = re.sub(r'\n?```\s*$', '', cleaned)
            try:
                parsed.append(json.loads(cleaned))
            except json.JSONDecodeError:
                parsed.append(content)
        final = json.dumps(parsed, indent=2, ensure_ascii=False)
    elif collect == 'concatenate':
        final = '\n\n'.join(content for _, content in results)
    else:  # raw
        final = results[-1][1] if results else ''

    if use_docx:
        docx_path = out.with_suffix('.docx')
        try:
            tools_dir = str(Path(__file__).parent / 'tools')
            if tools_dir not in sys.path:
                sys.path.insert(0, tools_dir)
            from md_to_docx import markdown_to_docx
            markdown_to_docx(final, str(docx_path))
            print(f"Wrote: {docx_path}")
        except ImportError:
            print("[WARN] md_to_docx not found; writing plain text instead.", file=sys.stderr)
            out.write_text(final, encoding='utf-8')
            print(f"Wrote: {out}")
    else:
        out.write_text(final, encoding='utf-8')
        print(f"Wrote: {out}")


# ── Execution engine ──────────────────────────────────────────────────────────

def _output_instructions(pf):
    """Extract plain text from the @output block (without the directive line and @end)."""
    if not pf.output_directive:
        return None
    text = re.sub(r'^@output(?:\s+\w+)?\s*\n?', '', pf.output_directive)
    text = re.sub(r'\n?@end\s*$', '', text)
    return text.strip()


def run_prompt_file(prompt_path, model=None, timeout=120, dry_run=False, debug=False,
                    use_docx=False):
    """Parse and execute a .prompt file. Returns list of (key, output) tuples."""
    pf = PromptFile(prompt_path)
    fm = pf.fm

    data_path = fm.get('data')
    output_path = fm.get('output')
    collect = fm.get('collect', 'raw')
    user_vars = fm.get('vars', {}) or {}
    sheet = fm.get('sheet')

    def _strval(v):
        if isinstance(v, bool):
            return 'true' if v else 'false'
        return str(v)

    base_ctx = {
        'DATE': date.today().isoformat(),
        **{str(k): _strval(v) for k, v in user_vars.items()},
    }

    output_instructions = _output_instructions(pf)

    # Load data
    data_items = []
    if data_path:
        full_path = pf.base_dir / data_path
        data_items = load_data(str(full_path), sheet=sheet)
        base_ctx['TOTAL'] = str(len(data_items))

    results = []

    # ── foreach mode ──
    if pf.foreach:
        if not data_items:
            print("[WARN] @foreach present but no data loaded.", file=sys.stderr)
            data_items = [('1', '', {'KEY': '1', 'VALUE': '', 'INDEX': '1'})]

        for idx, (key, value, row) in enumerate(data_items):
            ctx = {
                **base_ctx,
                'INDEX': str(idx + 1),
                'TOTAL': str(len(data_items)),
                'KEY': key,
                'VALUE': value,
                **row,
            }
            template = process_conditionals(pf.foreach, ctx)
            template = resolve_attachments(template, pf.base_dir)
            prompt = interpolate(template, ctx)
            if output_instructions:
                prompt = f"{prompt}\n\n{output_instructions}"

            print(f"  [{idx+1}/{len(data_items)}] {key[:60]}...")
            output = call_claude(prompt, system=pf.system, model=model, timeout=timeout,
                                 debug=debug, dry_run=dry_run)
            results.append((key, output))

    # ── pipeline / simple mode ──
    elif pf.steps:
        ctx = dict(base_ctx)
        prev_output = ''

        for step_idx, (step_name, template) in enumerate(pf.steps):
            ctx['PREV_OUTPUT'] = prev_output
            step_template = process_conditionals(template, ctx)
            step_template = resolve_attachments(step_template, pf.base_dir)
            prompt = interpolate(step_template, ctx)

            # Append output instructions only on the last step
            is_last = (step_idx == len(pf.steps) - 1)
            if output_instructions and is_last:
                prompt = f"{prompt}\n\n{output_instructions}"

            print(f"  Step: {step_name}...")
            output = call_claude(prompt, system=pf.system, model=model, timeout=timeout,
                                 debug=debug, dry_run=dry_run)

            ctx[f'step.{step_name}'] = output
            prev_output = output

        results = [('result', prev_output)]

    if results:
        write_output(results, output_path, collect, use_docx)

    return results


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Execute .prompt DSL files through Claude',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 run_claude.py examples/01_hello.prompt
  python3 run_claude.py examples/02_summarize.prompt --dry-run
  python3 run_claude.py examples/03_pipeline.prompt --model claude-opus-4-5
  python3 run_claude.py examples/07_xlsx_data.prompt --docx
""",
    )
    parser.add_argument('prompt_file', help='Path to .prompt file')
    parser.add_argument('--model', '-m', default=None,
                        help='Claude model (e.g. claude-opus-4-5)')
    parser.add_argument('--timeout', '-t', type=int, default=120,
                        help='Per-call timeout in seconds (default: 120)')
    parser.add_argument('--dry-run', '-n', action='store_true',
                        help='Preview prompts without calling Claude')
    parser.add_argument('--debug', '-d', action='store_true',
                        help='Show debug output')
    parser.add_argument('--docx', action='store_true',
                        help='Convert output to a Word document (.docx)')

    args = parser.parse_args()

    prompt_path = Path(args.prompt_file)
    if not prompt_path.exists():
        print(f"[ERROR] File not found: {args.prompt_file}", file=sys.stderr)
        sys.exit(1)

    print(f"Running: {prompt_path.name}")
    run_prompt_file(
        prompt_path,
        model=args.model,
        timeout=args.timeout,
        dry_run=args.dry_run,
        debug=args.debug,
        use_docx=args.docx,
    )


if __name__ == '__main__':
    main()
