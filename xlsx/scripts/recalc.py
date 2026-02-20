"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using Microsoft Excel.
macOS: AppleScript via osascript
Windows: COM automation via pywin32
"""

import json
import subprocess
import sys
from pathlib import Path

IS_WINDOWS = sys.platform == "win32"
IS_MACOS = sys.platform == "darwin"

if IS_WINDOWS:
    try:
        import win32com.client
    except ImportError:
        win32com = None

from openpyxl import load_workbook


def _recalc_macos(abs_path: str, timeout: int) -> dict | None:
    script = f'''
    tell application "Microsoft Excel"
        open POSIX file "{abs_path}"
        set theWorkbook to active workbook
        calculate theWorkbook
        save theWorkbook
        close theWorkbook
    end tell
    '''

    try:
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True,
            text=True,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        return {"error": "Excel timed out during recalculation"}

    if result.returncode != 0:
        return {"error": f"Excel recalculation failed: {result.stderr.strip()}"}

    return None


def _recalc_win32(abs_path: str) -> dict | None:
    if win32com is None:
        return {
            "error": (
                "pywin32 is required for Office automation on Windows. "
                "Install it with: pip install pywin32"
            )
        }

    app = win32com.client.Dispatch("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    workbook = None
    try:
        workbook = app.Workbooks.Open(abs_path)
        app.Calculate()
        workbook.Save()
    except Exception as e:
        return {"error": f"Excel recalculation failed: {e}"}
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)

    return None


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).resolve())

    if IS_WINDOWS:
        error = _recalc_win32(abs_path)
    elif IS_MACOS:
        error = _recalc_macos(abs_path, timeout)
    else:
        return {"error": f"Unsupported platform: {sys.platform}"}

    if error is not None:
        return error

    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count

        return result

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using Microsoft Excel")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
